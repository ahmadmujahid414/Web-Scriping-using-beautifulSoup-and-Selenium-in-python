import Edmunds_Automation as Ea
import webscraping as ws
import time
from openpyxl import Workbook

wb= Workbook()
sh1=wb.active
#time.sleep method is used many time because 
#to show clear automation
#I have automatic the script for all search result
#but in the main class I have done it just for one
#to save your time
#other wise I would have used a loop 
#to show you the result and data of all the seached results
sh1.cell(row=1,column=1,value="Names")
sh1.cell(row=1,column=2,value="Prices")
sh1.cell(row=1,column=3,value="VIN")
sh1.cell(row=1,column=4,value="Vehicle Summary ")
sh1.cell(row=1,column=5,value="Features & Specs")


user_zipcode = "23450"
user_radius = "100"

#below statment search through zip code 
Ea.search_postalCode(user_zipcode)
time.sleep(3)
#below statment search through radiu
Ea.search_radius(user_radius)
time.sleep(3)
#get access to the link with car names
ws.access_to_car_names()
time.sleep(2)
#return a list which have links with cars name 
Car = ws.Car_List()
time.sleep(3)
#show up the first car from the searched result
Ea.show_cars(Car[0]) 
time.sleep(3)  
#below lines of code
#store car name,price,vin,summary,features in excel file
sh1.cell(row=2,column=1,value=Car[0])
sh1.cell(row=2,column=2,value=ws.price())
sh1.cell(row=2,column=3,value=ws.VIN())
sh1.cell(row=2,column=4,value=ws.VehicleSummary())
sh1.cell(row=2,column=5,value=ws.Top_Features())
time.sleep(5)

Ea.back_to_main_page() #return to the main page from a specific car window
time.sleep(3)

wb.save("Vehicle_Details.xlsx")
Ea.quit()




